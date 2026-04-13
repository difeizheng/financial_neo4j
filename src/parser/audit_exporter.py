"""
audit_exporter.py

Exports an Excel audit workbook from a coverage report.
One worksheet per original sheet, color-coded by extraction status.

Columns: 行号 | 指标名称 | 公式 | 第1年值 | 状态 | 跳过原因 | 断裂依赖
Colors:
  green  (#92D050) — extracted, no broken deps
  orange (#FFC000) — extracted but has broken deps
  yellow (#FFFF00) — skipped (expected: header_row or skip_pattern)
  red    (#FF0000) — skipped (unexpected: not_meaningful_name or unknown)
"""

import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_FILL_GREEN  = PatternFill("solid", fgColor="FF92D050")
_FILL_ORANGE = PatternFill("solid", fgColor="FFFFC000")
_FILL_YELLOW = PatternFill("solid", fgColor="FFFFFF00")
_FILL_RED    = PatternFill("solid", fgColor="FFFF6666")
_FILL_HEADER = PatternFill("solid", fgColor="FF4472C4")

_FONT_HEADER = Font(bold=True, color="FFFFFFFF")
_FONT_NORMAL = Font(name="微软雅黑", size=10)

_HEADERS = ["行号", "指标名称", "公式", "第1年值", "状态", "跳过原因", "断裂依赖"]
_COL_WIDTHS = [8, 30, 40, 14, 10, 30, 50]


def export_audit_workbook(coverage: dict, indicators: list[dict]) -> bytes:
    """
    Build an openpyxl Workbook from the coverage report.
    Returns bytes suitable for st.download_button.
    """
    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    ind_lookup = _build_indicator_lookup(indicators)
    broken_lookup = _build_broken_dep_lookup(coverage.get("broken_dependencies", []))

    # Summary sheet first
    _write_summary_sheet(wb, coverage)

    for sheet_name, sheet_data in coverage["sheets"].items():
        _write_sheet(wb, sheet_name, sheet_data, ind_lookup, broken_lookup)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_summary_sheet(wb: Workbook, coverage: dict) -> None:
    ws = wb.create_sheet("覆盖率摘要")
    summary = coverage["summary"]

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    rows = [
        ("总内容行数", summary["total_content_rows"]),
        ("已提取行数", summary["extracted_rows"]),
        ("覆盖率", f"{summary['coverage_pct']:.1%}"),
        ("断裂依赖数", summary["broken_deps"]),
        ("", ""),
        ("工作表", "覆盖率"),
    ]
    for r in rows:
        ws.append(list(r))

    # Per-sheet summary rows
    for sheet_name, sd in coverage["sheets"].items():
        ws.append([sheet_name, f"{sd['coverage_pct']:.1%} ({sd['extracted']}/{sd['total']})"])

    # Style header row (row 6)
    for cell in ws[6]:
        cell.fill = _FILL_HEADER
        cell.font = _FONT_HEADER


def _write_sheet(
    wb: Workbook,
    sheet_name: str,
    sheet_data: dict,
    ind_lookup: dict,
    broken_lookup: dict,
) -> None:
    # Truncate sheet name to 31 chars (Excel limit)
    safe_name = sheet_name[:31]
    ws = wb.create_sheet(safe_name)

    # Header row
    for col_idx, (header, width) in enumerate(zip(_HEADERS, _COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _FILL_HEADER
        cell.font = _FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    for data_row in sheet_data["rows"]:
        row_num = data_row["row"]
        name = data_row["name"]
        status = data_row["status"]
        reason = data_row.get("reason", "")

        ind = ind_lookup.get((sheet_name, row_num))
        formula = ind["formula_raw"] if ind and ind.get("formula_raw") else ""
        value = ind["value_year1"] if ind and ind.get("value_year1") is not None else ""

        broken_list = broken_lookup.get((sheet_name, row_num), [])
        broken_str = "; ".join(
            f'{d["ref"]}→{d["target_sheet"]}行{d["target_row"]}' for d in broken_list
        )

        status_label = "已提取" if status == "extracted" else "已跳过"
        reason_label = _reason_label(reason)

        # Choose fill color
        if status == "extracted":
            fill = _FILL_ORANGE if broken_list else _FILL_GREEN
        else:
            fill = _FILL_YELLOW if reason in ("header_row",) or reason.startswith("skip_pattern:") else _FILL_RED

        values = [row_num, name, formula, value, status_label, reason_label, broken_str]
        excel_row = ws.max_row + 1
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.fill = fill
            cell.font = _FONT_NORMAL
            cell.alignment = Alignment(wrap_text=(col_idx in (3, 7)), vertical="top")


def _reason_label(reason: str) -> str:
    if not reason:
        return ""
    if reason == "header_row":
        return "表头行"
    if reason.startswith("skip_pattern:"):
        pat = reason[len("skip_pattern:"):]
        return f"跳过模式匹配: {pat}"
    if reason == "not_meaningful_name":
        return "名称无中文字符"
    if reason == "unknown":
        return "未知（可能是name_col配置错误）"
    return reason


def _build_indicator_lookup(indicators: list[dict]) -> dict:
    """Returns {(sheet, row): indicator_dict}."""
    return {(ind["sheet"], ind["row"]): ind for ind in indicators}


def _build_broken_dep_lookup(broken_deps: list[dict]) -> dict:
    """Returns {(source_sheet, source_name): [dep_dict, ...]}."""
    result: dict = {}
    for dep in broken_deps:
        key = (dep["source_sheet"], dep["source_name"])
        result.setdefault(key, []).append(dep)
    return result
