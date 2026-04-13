"""
excel_analyzer.py

Extracts structural metadata from an Excel file for LLM analysis.
Produces a compact representation of sheet layouts without loading all data.
"""
from __future__ import annotations
import logging
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

_MAX_HEADER_ROWS = 5
_MAX_SAMPLE_ROWS = 10
_MAX_FORMULA_EXAMPLES = 5
_MAX_COLS_PREVIEW = 15


def analyze_excel(excel_path: Path) -> dict:
    """
    Extract structural metadata from an Excel file.

    Returns a dict with a 'sheets' list, each containing:
      - name, max_row, max_col
      - headers: first _MAX_HEADER_ROWS rows (all columns up to _MAX_COLS_PREVIEW)
      - sample_rows: rows after headers (up to _MAX_SAMPLE_ROWS)
      - formula_examples: up to _MAX_FORMULA_EXAMPLES cells with formulas
      - merged_cells: list of merged cell range strings
    """
    logger.info(f"Analyzing Excel structure: {excel_path}")
    wb = load_workbook(excel_path, data_only=False, read_only=False)

    sheets_meta = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        # Limit columns to preview
        preview_cols = min(max_col, _MAX_COLS_PREVIEW)

        def read_row(row_num: int) -> dict:
            cells = {}
            for col in range(1, preview_cols + 1):
                val = ws.cell(row=row_num, column=col).value
                if val is not None:
                    cells[get_column_letter(col)] = str(val)[:80]
            return {"row": row_num, "cells": cells}

        # Header rows
        headers = []
        for r in range(1, min(_MAX_HEADER_ROWS + 1, max_row + 1)):
            row_data = read_row(r)
            if row_data["cells"]:
                headers.append(row_data)

        # Sample rows (after headers)
        sample_rows = []
        start = _MAX_HEADER_ROWS + 1
        count = 0
        for r in range(start, max_row + 1):
            if count >= _MAX_SAMPLE_ROWS:
                break
            row_data = read_row(r)
            if row_data["cells"]:
                sample_rows.append(row_data)
                count += 1

        # Formula examples
        formula_examples = []
        for row in ws.iter_rows(max_row=min(max_row, 100)):
            for cell in row:
                if len(formula_examples) >= _MAX_FORMULA_EXAMPLES:
                    break
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_examples.append({
                        "row": cell.row,
                        "col": get_column_letter(cell.column),
                        "formula": cell.value[:100],
                    })
            if len(formula_examples) >= _MAX_FORMULA_EXAMPLES:
                break

        # Merged cells
        merged = [str(r) for r in ws.merged_cells.ranges][:20]

        sheets_meta.append({
            "name": sheet_name,
            "max_row": max_row,
            "max_col": max_col,
            "headers": headers,
            "sample_rows": sample_rows,
            "formula_examples": formula_examples,
            "merged_cells": merged,
        })

    wb.close()
    logger.info(f"Analyzed {len(sheets_meta)} sheets")
    return {"sheets": sheets_meta}
