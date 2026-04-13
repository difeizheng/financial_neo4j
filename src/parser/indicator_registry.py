"""
indicator_registry.py

Extracts financial indicators from all sheets of the Excel model.
Each indicator becomes a node in the Neo4j graph.

Key challenge: the parameter sheet (参数输入表) has "yearly expansion" rows
where a single indicator (e.g. "年度达产率") is followed by 48 rows of
"合作期第N年" values. These must be collapsed into one indicator node.

The extractor also tracks parent-child relationships via section numbers:
- Rows with non-empty section number (like "1", "2", "15") are parents
- Rows with empty section number or decimal like "1.1" are children
- Children get a CHILD_OF relationship edge to their parent
"""

import re
import json
import logging
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from src.parser.sheet_config import SHEET_CONFIGS, SHEET_CATEGORIES
from typing import Callable, Optional

logger = logging.getLogger(__name__)

# Regex: detect yearly expansion rows like "合作期第1年", "建设期第3年"
_YEARLY_ROW_RE = re.compile(r"(合作期|建设期)第\d+[年月]?")

# Regex: detect pure numeric or section-number strings like "1", "1.1", "一", "（一）"
_SECTION_NUM_RE = re.compile(r"^[\d一二三四五六七八九十（）\.\s]+$")

# Regex: detect decimal section numbers (children)
_DECIMAL_SECTION_RE = re.compile(r"^\d+\.\d+$")

# Regex: detect pure numeric string (potential Excel date serial)
_NUMERIC_RE = re.compile(r"^-?\d+(\.\d+)?$")

# Regex: detect Chinese numerals (should be kept as-is, not converted)
_CHINESE_NUM_RE = re.compile(r"^[一二三四五六七八九十]+$")


def _col_idx(letter: str) -> int:
    """Convert column letter to 1-based index."""
    return column_index_from_string(letter)


def _excel_serial_to_year_month(serial: int) -> str:
    """Convert Excel date serial number to 'yyyy年mm月' format.

    Excel serial: 1 = January 1, 1900 (Windows) or January 1, 1904 (Mac)
    We use Windows Excel 1900 date system.
    """
    if serial is None:
        return None
    try:
        # Excel base date: January 1, 1900
        base_date = datetime(1900, 1, 1)
        # Excel has a bug where it treats 1900 as a leap year
        # So we add 1 day to account for the fictional Feb 29, 1900
        target_date = base_date + timedelta(days=int(serial) - 1)
        return f"{target_date.year}年{target_date.month:02d}月"
    except (ValueError, TypeError):
        return str(serial)


def _cell_val(ws, row: int, col_letter: str):
    """Return stripped string value of a cell, or None."""
    if col_letter is None:
        return None
    val = ws.cell(row=row, column=_col_idx(col_letter)).value
    if val is None:
        return None
    return str(val).strip()


def _is_yearly_expansion(name: str) -> bool:
    return bool(_YEARLY_ROW_RE.search(name))


def _is_parent_section(section_num: str) -> bool:
    """Return True if section_num indicates a parent (non-decimal, non-empty)."""
    if not section_num:
        return False
    # Decimal like "1.1", "15.1" are children, not parents
    if _DECIMAL_SECTION_RE.match(section_num.strip()):
        return False
    # Pure Chinese numeral like "一" is kept as-is (not a parent number)
    # but we still treat it as a parent for hierarchy tracking
    return True


def _is_meaningful_name(name: str) -> bool:
    """Return True if the string looks like a real indicator name (not a header/number)."""
    if not name:
        return False
    # Must contain at least one Chinese character
    if not re.search(r"[\u4e00-\u9fff]", name):
        return False
    return True


def _make_id(sheet_name: str, name: str, row: int) -> str:
    """Create a unique indicator ID."""
    # Sanitize: remove special chars
    safe_sheet = re.sub(r"[^\w\u4e00-\u9fff]", "_", sheet_name)
    safe_name = re.sub(r"[^\w\u4e00-\u9fff]", "_", name)
    return f"{safe_sheet}__{safe_name}__{row}"


def extract_indicators(
    excel_path: Path,
    sheet_configs: Optional[dict] = None,
    sheet_categories: Optional[dict] = None,
    progress_callback: Optional[Callable[[str, float], None]] = None,
) -> tuple[list[dict], list[dict]]:
    """
    Parse the Excel file and return a tuple of (indicators, CHILD_OF edges).

    sheet_configs: if None, uses the hardcoded SHEET_CONFIGS (CLI path).
    sheet_categories: if None, uses the hardcoded SHEET_CATEGORIES.
    progress_callback: optional callable(message, percent) for progress reporting.

    Returns:
        tuple: (indicators list, CHILD_OF edges list)
    """
    configs = sheet_configs if sheet_configs is not None else SHEET_CONFIGS
    categories = sheet_categories if sheet_categories is not None else SHEET_CATEGORIES

    logger.info(f"Loading workbook (formulas): {excel_path}")
    wb = load_workbook(excel_path, data_only=False, read_only=True)

    all_indicators = []
    all_edges = []  # CHILD_OF edges

    sheet_names = list(configs.keys())
    total = len(sheet_names)

    for i, sheet_name in enumerate(sheet_names):
        if sheet_name not in wb.sheetnames:
            logger.warning(f"Sheet not found: {sheet_name!r}")
            continue

        ws = wb[sheet_name]
        logger.info(f"Processing sheet: {sheet_name}")
        if progress_callback:
            progress_callback(f"解析工作表: {sheet_name}", i / total)

        cfg = configs[sheet_name]
        indicators, edges = _extract_from_sheet(ws, sheet_name, cfg, categories)
        logger.info(f"  → {len(indicators)} indicators, {len(edges)} CHILD_OF edges")
        all_indicators.extend(indicators)
        all_edges.extend(edges)

    wb.close()
    logger.info(f"Total indicators extracted: {len(all_indicators)}")
    logger.info(f"Total CHILD_OF edges: {len(all_edges)}")
    return all_indicators, all_edges


def _extract_from_sheet(ws, sheet_name: str, cfg: dict, sheet_categories: dict) -> tuple[list[dict], list[dict]]:
    """Extract indicators from a single sheet.

    Returns:
        tuple: (indicators list, CHILD_OF edges list)
    """
    name_col = cfg["name_col"]
    formula_col = cfg["formula_col"]
    unit_col = cfg.get("unit_col")
    number_col = cfg.get("number_col")
    header_rows = set(cfg.get("header_rows", []))
    skip_patterns = cfg.get("skip_patterns", [])  # Only skip "下拉菜单" etc., not yearly rows
    is_input = cfg.get("is_input", False)
    sheet_category = sheet_categories.get(sheet_name, "其他")

    indicators = []
    edges = []

    current_category = ""  # tracks section headers in 参数输入表
    current_parent = None  # {id, name, section_number} for hierarchy tracking
    child_counter = {}    # {parent_section: count} for generating "1-1", "1-2" etc.

    # Track last meaningful name for numeric row association
    _last_meaningful_name = None
    _last_meaningful_row = None
    _last_section_number = None

    max_row = ws.max_row or 0

    for row_num in range(1, max_row + 1):
        if row_num in header_rows:
            continue

        name = _cell_val(ws, row_num, name_col)
        if not name:
            # Reset parent when name is empty (end of a section)
            current_parent = None
            continue

        # Skip only non-yearly patterns (e.g., "下拉菜单")
        # Do NOT skip "合作期第", "建设期第" - they are yearly expansion rows
        if any(pat in name for pat in skip_patterns) and not _YEARLY_ROW_RE.search(name):
            continue

        # Track category headers (参数输入表 uses column B for section headers)
        if cfg.get("category_col"):
            cat_val = _cell_val(ws, row_num, cfg["category_col"])
            if cat_val and re.search(r"[\u4e00-\u9fff]", cat_val) and len(cat_val) < 20:
                current_category = cat_val

        # Get section number
        section_number = _cell_val(ws, row_num, number_col) if number_col else None

        # Track meaningful names for potential parent-child relationship with numeric rows
        if _is_meaningful_name(name):
            _last_meaningful_name = name
            _last_meaningful_row = row_num
            _last_section_number = section_number

        # Determine if this is a yearly expansion row (合作期第n年 / 建设期第n月)
        yearly_match = _YEARLY_ROW_RE.search(name)

        # Determine if this is a numeric-only name (potential Excel date serial)
        is_numeric_name = bool(_NUMERIC_RE.match(name))

        # Check for parent-child relationship
        is_parent = _is_parent_section(section_number)
        is_decimal_child = _DECIMAL_SECTION_RE.match(section_number.strip()) if section_number else False

        # Handle yearly expansion row (合作期第n年 / 建设期第n月)
        # These are children of the current root parent
        if yearly_match:
            # Find the root parent (not the immediate parent, but the top-level parent)
            root_parent = current_parent
            if root_parent:
                # Create child indicator with parent prefix
                child_name = f"{root_parent['name']}_{name}"
                indicator = _create_indicator(
                    sheet_name, sheet_category, current_category,
                    row_num, child_name, name_col, formula_col, unit_col,
                    ws, formula_col, is_input,
                    parent_id=root_parent["id"],
                    parent_name=root_parent["name"],
                    section_number=f"{root_parent['section_number']}-{child_counter.get(root_parent['section_number'], 1)}"
                )
                # Increment child counter for root parent
                child_counter[root_parent['section_number']] = child_counter.get(root_parent['section_number'], 1) + 1

                # Create CHILD_OF edge
                edges.append({
                    "source_id": indicator["id"],
                    "target_id": root_parent["id"],
                    "relationship": "CHILD_OF",
                })

                indicators.append(indicator)
            continue

        # Handle numeric name (Excel date serial) - convert to year-month format
        # These become children of the current parent (if exists)
        # Also check if previous row was a meaningful name (potential parent for numeric rows)
        if is_numeric_name and not _is_meaningful_name(name):
            # Look for parent in hierarchy: use current_parent OR
            # check if previous meaningful name could be parent
            parent = current_parent

            # If no current_parent but there was a previous meaningful row that could be parent
            if not parent and _last_meaningful_name:
                # Create a pseudo-parent for numeric rows
                parent = {
                    "id": _make_id(sheet_name, _last_meaningful_name, _last_meaningful_row),
                    "name": _last_meaningful_name,
                    "section_number": _last_section_number or "unknown",
                }

            if parent:
                try:
                    serial_num = int(float(name))
                    year_month = _excel_serial_to_year_month(serial_num)
                    if year_month:
                        child_name = f"{parent['name']}_{year_month}"
                        indicator = _create_indicator(
                            sheet_name, sheet_category, current_category,
                            row_num, child_name, name_col, formula_col, unit_col,
                            ws, formula_col, is_input,
                            parent_id=parent["id"],
                            parent_name=parent["name"],
                            section_number=f"{parent['section_number']}-{child_counter.get(parent['section_number'], 1)}"
                        )
                        section_key = parent['section_number']
                        child_counter[section_key] = child_counter.get(section_key, 1) + 1

                        edges.append({
                            "source_id": indicator["id"],
                            "target_id": parent["id"],
                            "relationship": "CHILD_OF",
                        })
                        indicators.append(indicator)
                except (ValueError, TypeError):
                    pass  # Not a valid number, skip
            continue

        # Skip if not a meaningful name
        if not _is_meaningful_name(name):
            continue

        # This is a regular indicator (parent or child)
        # Get formula / value
        formula_raw = _cell_val(ws, row_num, formula_col) if formula_col else None

        # Get unit
        unit = _cell_val(ws, row_num, unit_col) if unit_col else None

        # Handle decimal section numbers (like "1.1") as children
        if is_decimal_child and current_parent:
            indicator = _create_indicator(
                sheet_name, sheet_category, current_category,
                row_num, name, name_col, formula_col, unit_col,
                ws, formula_col, is_input,
                parent_id=current_parent["id"],
                parent_name=current_parent["name"],
                section_number=section_number
            )

            edges.append({
                "source_id": indicator["id"],
                "target_id": current_parent["id"],
                "relationship": "CHILD_OF",
            })
            indicators.append(indicator)
            continue

        # Handle empty section number (child of current parent)
        if not section_number and current_parent:
            # Generate child number like "1-1", "1-2" etc.
            child_num = child_counter.get(current_parent['section_number'], 1)
            generated_section = f"{current_parent['section_number']}-{child_num}"
            child_counter[current_parent['section_number']] = child_num + 1

            indicator = _create_indicator(
                sheet_name, sheet_category, current_category,
                row_num, name, name_col, formula_col, unit_col,
                ws, formula_col, is_input,
                parent_id=current_parent["id"],
                parent_name=current_parent["name"],
                section_number=generated_section
            )

            edges.append({
                "source_id": indicator["id"],
                "target_id": current_parent["id"],
                "relationship": "CHILD_OF",
            })
            indicators.append(indicator)
            continue

        # This is a parent indicator (has non-empty, non-decimal section number)
        indicator = _create_indicator(
            sheet_name, sheet_category, current_category,
            row_num, name, name_col, formula_col, unit_col,
            ws, formula_col, is_input,
            parent_id=None,
            parent_name=None,
            section_number=section_number
        )
        indicators.append(indicator)

        # Update current parent
        current_parent = {
            "id": indicator["id"],
            "name": indicator["name"],
            "section_number": section_number or "",
        }
        # Initialize child counter for this parent
        if section_number:
            child_counter[section_number] = 1

    return indicators, edges


def _create_indicator(
    sheet_name: str,
    sheet_category: str,
    category: str,
    row_num: int,
    name: str,
    name_col: str,
    formula_col: str,
    unit_col: str,
    ws,
    formula_col_for_value: str,
    is_input: bool,
    parent_id: str = None,
    parent_name: str = None,
    section_number: str = None,
) -> dict:
    """Create an indicator dict with all fields."""
    formula_raw = _cell_val(ws, row_num, formula_col) if formula_col else None
    unit = _cell_val(ws, row_num, unit_col) if unit_col else None

    indicator = {
        "id": _make_id(sheet_name, name, row_num),
        "name": name,
        "sheet": sheet_name,
        "sheet_category": sheet_category,
        "category": category,
        "row": row_num,
        "formula_raw": formula_raw,
        "unit": unit or "万元",
        "section_number": section_number,
        "is_input": is_input,
        "is_circular": False,  # will be updated by formula_parser
        "parent_id": parent_id,
        "parent_name": parent_name,
    }
    return indicator


def save_indicators(indicators: list[dict], output_path: Path) -> None:
    output_path.parent.mkdir(exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(indicators, f, ensure_ascii=False, indent=2)
    logger.info(f"Saved {len(indicators)} indicators to {output_path}")


def save_child_of_edges(edges: list[dict], output_path: Path) -> None:
    """Save CHILD_OF edges to JSON file."""
    output_path.parent.mkdir(exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(edges, f, ensure_ascii=False, indent=2)
    logger.info(f"Saved {len(edges)} CHILD_OF edges to {output_path}")


def load_indicators(path: Path) -> list[dict]:
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def load_child_of_edges(path: Path) -> list[dict]:
    """Load CHILD_OF edges from JSON file."""
    with open(path, encoding="utf-8") as f:
        return json.load(f)
