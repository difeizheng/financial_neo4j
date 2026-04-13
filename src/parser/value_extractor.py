"""
value_extractor.py

Loads the Excel workbook in data_only mode to extract computed values.
Enriches each indicator with:
  - value_year1: first operational year value (column F, or first non-zero)
  - values_json: full time series as JSON string (columns F through end)
"""

import json
import logging
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from src.parser.sheet_config import SHEET_CONFIGS
from typing import Optional

logger = logging.getLogger(__name__)

# First data column (year 1 of cooperation period) is typically column F
_DEFAULT_FIRST_DATA_COL = "F"
# Max number of yearly columns to extract (48 years + some buffer)
_MAX_YEARS = 55


def _col_idx(letter: str) -> int:
    return column_index_from_string(letter)


def extract_values(
    excel_path: Path,
    indicators: list[dict],
    sheet_configs: Optional[dict] = None,
    progress_callback=None,
    stop_check=None,
) -> list[dict]:
    """
    For each indicator, extract its computed value(s) from the Excel.
    Modifies indicators in-place and returns them.

    sheet_configs: if None, uses the hardcoded SHEET_CONFIGS (CLI path).
    progress_callback: optional callable(message, percent 0-1) for per-sheet progress.
    stop_check: optional callable() that raises an exception to abort early.
    """
    configs = sheet_configs if sheet_configs is not None else SHEET_CONFIGS

    logger.info(f"Loading workbook (data_only): {excel_path}")
    wb = load_workbook(excel_path, data_only=True, read_only=True)

    # Build lookup: (sheet, row) -> indicator index in list
    lookup = {}
    for i, ind in enumerate(indicators):
        lookup[(ind["sheet"], ind["row"])] = i

    valid_sheets = [s for s in configs if s in wb.sheetnames]
    total = len(valid_sheets)

    for sheet_idx, sheet_name in enumerate(valid_sheets):
        if stop_check:
            stop_check()
        if progress_callback:
            progress_callback(f"提取数值: {sheet_name} ({sheet_idx + 1}/{total})", sheet_idx / total)

        cfg = configs[sheet_name]
        ws = wb[sheet_name]
        formula_col = cfg.get("formula_col") or _DEFAULT_FIRST_DATA_COL
        first_col_idx = _col_idx(formula_col)
        last_col_idx = first_col_idx + _MAX_YEARS - 1

        # Stream rows in one pass — avoids O(N²) random seeks in read_only mode.
        # In read_only mode openpyxl may return EmptyCell objects (no .row attr)
        # for sparse rows; find the first real cell to get the row number.
        for row_cells in ws.iter_rows(min_col=1, max_col=last_col_idx):
            row_num = next((c.row for c in row_cells if hasattr(c, "row")), None)
            if row_num is None:
                continue  # completely empty row, skip
            key = (sheet_name, row_num)
            if key not in lookup:
                continue

            idx = lookup[key]
            values = []
            for col_offset in range(_MAX_YEARS):
                tuple_idx = first_col_idx + col_offset - 1  # 0-based (min_col=1)
                cell_val = row_cells[tuple_idx].value if tuple_idx < len(row_cells) else None
                if cell_val is None:
                    values.append(None)
                else:
                    try:
                        values.append(float(cell_val))
                    except (TypeError, ValueError):
                        values.append(str(cell_val))

            value_year1 = next((v for v in values if v is not None and v != 0 and v != ""), None)
            indicators[idx]["value_year1"] = value_year1
            indicators[idx]["values_json"] = json.dumps(values, ensure_ascii=False)

    wb.close()
    logger.info("Value extraction complete")
    return indicators
