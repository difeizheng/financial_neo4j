"""
coverage_scanner.py

Compares the original Excel against extracted indicators to detect:
  1. Rows with content that were silently skipped (and why)
  2. Formula references that point to rows not extracted as indicators

Reuses private symbols from formula_parser and indicator_registry to ensure
the skip-reason classification exactly mirrors the actual extraction logic.
"""

import json
import logging
import time as _time
from pathlib import Path
from openpyxl import load_workbook

from src.parser.formula_parser import (
    _CROSS_SHEET_REF,
    _SAME_SHEET_REF,
    _normalize_sheet_name,
    build_row_index,
)
from src.parser.indicator_registry import _is_meaningful_name, _cell_val

logger = logging.getLogger(__name__)


# ── Public API ─────────────────────────────────────────────────────────────────

def scan_coverage(
    excel_path: Path,
    sheet_configs: dict,
    indicators: list[dict],
) -> dict:
    """
    Open the original Excel once and produce a full coverage report.

    Returns:
    {
      "summary": {
        "total_content_rows": int,   # rows with non-empty name_col across all sheets
        "extracted_rows": int,
        "coverage_pct": float,       # 0.0–1.0
        "broken_deps": int,
      },
      "sheets": {
        "SheetName": {
          "total": int, "extracted": int, "coverage_pct": float,
          "rows": [
            {"row": int, "name": str, "status": "extracted"|"skipped", "reason": str}
          ]
        }
      },
      "broken_dependencies": [
        {
          "source_name": str, "source_sheet": str,
          "formula": str,      # truncated to 200 chars
          "ref": str,          # matched cell reference fragment
          "target_sheet": str, "target_row": int,
        }
      ]
    }
    """
    logger.info(f"Starting coverage scan: {excel_path.name}")
    t0 = _time.time()

    wb = load_workbook(excel_path, data_only=False, read_only=True)

    # Build lookup: (sheet, row) -> indicator_id
    row_index = build_row_index(indicators)
    # Build per-sheet set of extracted row numbers for fast lookup
    extracted_by_sheet: dict[str, set[int]] = {}
    for ind in indicators:
        extracted_by_sheet.setdefault(ind["sheet"], set()).add(ind["row"])

    sheets_report: dict[str, dict] = {}
    total_content = 0
    total_extracted = 0

    for sheet_name, cfg in sheet_configs.items():
        if sheet_name not in wb.sheetnames:
            logger.warning(f"Sheet not in workbook: {sheet_name!r}")
            continue

        ws = wb[sheet_name]
        extracted_rows = extracted_by_sheet.get(sheet_name, set())
        rows = _scan_sheet_rows(ws, sheet_name, cfg, extracted_rows)

        n_total = len(rows)
        n_extracted = sum(1 for r in rows if r["status"] == "extracted")
        total_content += n_total
        total_extracted += n_extracted

        sheets_report[sheet_name] = {
            "total": n_total,
            "extracted": n_extracted,
            "coverage_pct": n_extracted / n_total if n_total else 1.0,
            "rows": rows,
        }

    wb.close()

    broken_deps = _audit_dependencies(indicators, row_index, sheet_configs)

    coverage_pct = total_extracted / total_content if total_content else 1.0
    report = {
        "summary": {
            "total_content_rows": total_content,
            "extracted_rows": total_extracted,
            "coverage_pct": coverage_pct,
            "broken_deps": len(broken_deps),
        },
        "sheets": sheets_report,
        "broken_dependencies": broken_deps,
    }

    elapsed = _time.time() - t0
    logger.info(
        f"Coverage scan done in {elapsed:.1f}s: "
        f"{total_extracted}/{total_content} rows ({coverage_pct:.1%}), "
        f"{len(broken_deps)} broken deps"
    )
    return report


def format_coverage_feedback(coverage: dict, threshold: float = 0.90) -> str:
    """
    Format the coverage report as a structured Chinese feedback string for
    passing to run_step1(feedback=...).

    Only meaningful when coverage_pct < threshold, but safe to call regardless.
    Each problem section lists at most 5 examples to keep the prompt concise.
    """
    summary = coverage["summary"]
    pct = summary["coverage_pct"]
    total = summary["total_content_rows"]
    extracted = summary["extracted_rows"]
    broken = summary["broken_deps"]

    lines = [
        "## 覆盖率验证报告（自动生成）",
        "",
        f"总体覆盖率：{pct:.1%}（{extracted}/{total}行），"
        + ("低于阈值，请根据以下问题调整配置。" if pct < threshold else "已达标，以下为参考信息。"),
        "",
    ]

    # Collect skipped rows by reason category
    pattern_rows: list[dict] = []
    header_rows_list: list[dict] = []
    not_meaningful_rows: list[dict] = []
    unknown_rows: list[dict] = []

    for sheet_name, sheet_data in coverage["sheets"].items():
        for row in sheet_data["rows"]:
            if row["status"] != "skipped":
                continue
            reason = row.get("reason", "")
            entry = {"sheet": sheet_name, "row": row["row"], "name": row["name"], "reason": reason}
            if reason.startswith("skip_pattern:"):
                pattern_rows.append(entry)
            elif reason == "header_row":
                header_rows_list.append(entry)
            elif reason == "not_meaningful_name":
                not_meaningful_rows.append(entry)
            else:
                unknown_rows.append(entry)

    # Section 1: skip_patterns over-matching
    if pattern_rows:
        lines.append("### 问题1：skip_patterns过度匹配")
        lines.append("以下行被skip_patterns跳过，但可能是有效指标：")
        for e in pattern_rows[:5]:
            pat = e["reason"].replace("skip_pattern:", "")
            lines.append(f'- 工作表「{e["sheet"]}」第{e["row"]}行：名称="{e["name"]}"，匹配模式="{pat}"')
        if len(pattern_rows) > 5:
            lines.append(f"  （另有 {len(pattern_rows) - 5} 行未列出）")
        lines.append('建议：检查上述工作表的skip_patterns，将过于宽泛的模式（如"第"）改为更精确的模式（如"合作期第"或"建设期第"）。')
        lines.append("")

    # Section 2: header_rows over-counting
    if header_rows_list:
        lines.append("### 问题2：header_rows配置过多，覆盖了真实数据行")
        lines.append("以下行被标记为表头行，但看起来包含真实指标：")
        for e in header_rows_list[:5]:
            lines.append(f'- 工作表「{e["sheet"]}」第{e["row"]}行：名称="{e["name"]}"')
        if len(header_rows_list) > 5:
            lines.append(f"  （另有 {len(header_rows_list) - 5} 行未列出）")
        lines.append("建议：减少上述工作表的header_rows范围。")
        lines.append("")

    # Section 3: not_meaningful_name (informational only — system limitation)
    if not_meaningful_rows:
        lines.append("### 问题3：纯英文/数字指标名被过滤（系统限制，仅供参考）")
        lines.append("以下行因名称不含中文字符被跳过，这是系统限制，无法通过配置修复：")
        for e in not_meaningful_rows[:5]:
            lines.append(f'- 工作表「{e["sheet"]}」第{e["row"]}行：名称="{e["name"]}"')
        if len(not_meaningful_rows) > 5:
            lines.append(f"  （另有 {len(not_meaningful_rows) - 5} 行未列出）")
        lines.append("")

    # Section 4: unknown skips (likely name_col misconfiguration)
    if unknown_rows:
        lines.append("### 问题4：未知原因跳过（可能是name_col配置错误或合并单元格）")
        lines.append("以下行通过了所有过滤条件但未被提取，可能是name_col指向了错误的列：")
        for e in unknown_rows[:5]:
            lines.append(f'- 工作表「{e["sheet"]}」第{e["row"]}行：名称="{e["name"]}"')
        if len(unknown_rows) > 5:
            lines.append(f"  （另有 {len(unknown_rows) - 5} 行未列出）")
        lines.append("建议：检查上述工作表的name_col是否正确。")
        lines.append("")

    # Section 5: broken dependencies
    broken_deps = coverage.get("broken_dependencies", [])
    if broken_deps:
        lines.append(f"### 问题5：断裂的公式依赖（共{len(broken_deps)}处）")
        lines.append("以下指标的公式引用了未被提取的行：")
        for dep in broken_deps[:5]:
            lines.append(
                f'- 「{dep["source_name"]}」（{dep["source_sheet"]}）：'
                f'公式引用了「{dep["target_sheet"]}」第{dep["target_row"]}行（未提取）'
            )
        if len(broken_deps) > 5:
            lines.append(f"  （另有 {len(broken_deps) - 5} 处未列出）")
        lines.append("修复上述跳过问题后，这些依赖应自动恢复。")
        lines.append("")

    lines.append("请重新生成配置，重点修正上述问题。")
    return "\n".join(lines)


def save_coverage(coverage: dict, output_path: Path) -> None:
    """Atomically write coverage.json (tmp → replace with retry for Windows)."""
    content = json.dumps(coverage, ensure_ascii=False, indent=2)
    tmp = output_path.with_suffix(".json.tmp")
    tmp.write_text(content, encoding="utf-8")
    for attempt in range(6):
        try:
            tmp.replace(output_path)
            return
        except PermissionError:
            if attempt < 5:
                _time.sleep(0.05)
            else:
                output_path.write_text(content, encoding="utf-8")
                try:
                    tmp.unlink(missing_ok=True)
                except OSError:
                    pass


def load_coverage(path: Path) -> dict | None:
    """Load coverage.json; return None if file does not exist or is corrupt."""
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None


# ── Internal helpers ───────────────────────────────────────────────────────────

def _scan_sheet_rows(
    ws,
    sheet_name: str,
    cfg: dict,
    extracted_rows: set[int],
) -> list[dict]:
    """
    Walk every row in ws. For each row where name_col is non-empty, classify it.

    Gate order mirrors _extract_from_sheet() exactly:
      1. header_rows check
      2. skip_patterns substring match
      3. _is_meaningful_name()
      4. presence in extracted_rows (should always be True if we reach here)
    """
    name_col = cfg["name_col"]
    header_rows = set(cfg.get("header_rows", []))
    skip_patterns = cfg.get("skip_patterns", [])

    max_row = ws.max_row or 0
    result = []

    for row_num in range(1, max_row + 1):
        name = _cell_val(ws, row_num, name_col)
        if not name:
            continue  # truly empty — not a candidate, don't report

        if row_num in header_rows:
            result.append({"row": row_num, "name": name, "status": "skipped", "reason": "header_row"})
            continue

        matched_pat = next((p for p in skip_patterns if p in name), None)
        if matched_pat is not None:
            result.append({"row": row_num, "name": name, "status": "skipped",
                           "reason": f"skip_pattern:{matched_pat}"})
            continue

        if not _is_meaningful_name(name):
            result.append({"row": row_num, "name": name, "status": "skipped",
                           "reason": "not_meaningful_name"})
            continue

        # Passed all filters
        if row_num in extracted_rows:
            result.append({"row": row_num, "name": name, "status": "extracted", "reason": ""})
        else:
            # Passed filters but not in extracted set — likely merged cell or name_col issue
            result.append({"row": row_num, "name": name, "status": "skipped", "reason": "unknown"})

    return result


def _audit_dependencies(
    indicators: list[dict],
    row_index: dict,
    sheet_configs: dict,
) -> list[dict]:
    """
    Re-parse every formula using the same two-pass regex approach as
    parse_dependencies(), and report references that point to rows not in row_index.
    """
    known_sheets = {ind["sheet"] for ind in indicators}
    broken: list[dict] = []
    seen: set[tuple] = set()  # (source_id, target_sheet, target_row)

    for ind in indicators:
        formula = ind.get("formula_raw") or ""
        if not formula or not formula.startswith("="):
            continue

        source_id = ind["id"]
        source_sheet = ind["sheet"]
        formula_trunc = formula[:200]

        # Pass 1: cross-sheet references
        for match in _CROSS_SHEET_REF.finditer(formula):
            raw_sheet, _col, row_str = match.groups()
            sheet_name = _normalize_sheet_name(raw_sheet)
            row_num = int(row_str)

            # Resolve sheet name (same fuzzy logic as parse_dependencies)
            resolved_sheet = sheet_name
            if (sheet_name, row_num) not in row_index:
                for known in known_sheets:
                    if sheet_name in known or known in sheet_name:
                        if (known, row_num) in row_index:
                            resolved_sheet = known
                            break

            key = (source_id, resolved_sheet, row_num)
            if key in seen:
                continue
            seen.add(key)

            if (resolved_sheet, row_num) not in row_index:
                broken.append({
                    "source_name": ind["name"],
                    "source_sheet": source_sheet,
                    "formula": formula_trunc,
                    "ref": match.group(0),
                    "target_sheet": resolved_sheet,
                    "target_row": row_num,
                })

        # Pass 2: same-sheet references (strip cross-sheet refs first)
        formula_stripped = _CROSS_SHEET_REF.sub("__REMOVED__", formula)
        for match in _SAME_SHEET_REF.finditer(formula_stripped):
            _col, row_str = match.groups()
            row_num = int(row_str)

            key = (source_id, source_sheet, row_num)
            if key in seen:
                continue
            seen.add(key)

            if (source_sheet, row_num) not in row_index:
                broken.append({
                    "source_name": ind["name"],
                    "source_sheet": source_sheet,
                    "formula": formula_trunc,
                    "ref": match.group(0),
                    "target_sheet": source_sheet,
                    "target_row": row_num,
                })

    return broken
